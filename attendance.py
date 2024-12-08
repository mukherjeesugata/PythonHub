import openpyxl
from openpyxl.styles import Alignment, Font

class AttendanceRecord:
    def __init__(self):
        self.records={}

    def add_records(self,student_id,student_name,attendance):
        if student_id in self.records:
            print(f"Students ID {student_id} already exists.")
        else:
            self.records[student_id]= { 
                "name":student_name,
                "attendance":attendance
            }
            print(f"Records added for {student_name}.")
            
    def view_records(self):
        if not self.records:
            print("No attendance records found.")
        else:
            print("\nAttendance Records")
            print("-" * 50)
            print(f"{'ID':<15}{'Name':<25}{'Attendance':<10}")
            print("-" * 50)
            for student_id, details in self.records.items():
                print(f"{student_id:<15}{details['name']:<25}{details['attendance']:<10}")
            print("-" * 50)

    def update_records(self,student_id,student_name,attendance):
        if student_id in self.records:
           if student_name:  # Update the name only if provided
               self.records[student_id]['name'] = student_name
           if attendance:  # Update the attendance only if provided
               self.records[student_id]['attendance'] = attendance
           print(f"Record for {student_id} updated successfully.")
        else:
           print(f"Student ID {student_id} not found. Please add the record first.")

    def replace_student_id(self,old_id,new_id):
        if old_id in self.records:
            self.records[new_id]=self.records[old_id]
            del self.records[old_id]
            print(f"Student ID replaced from {old_id} to {new_id}.")
        else:
            print(f"Student ID {old_id} not found.")

    def delete_records(self,student_id):
        if student_id in self.records:
            student_name = self.records[student_id]['name']
            del self.records[student_id]
            print(f"Record for {student_name} deleted.") 
        else:
            print("Student ID is not found.")

    def update_excel(self):
        wb = openpyxl.Workbook()
        sheet = wb.active 
        sheet.title = "Attendance Records"

        headers = ["Student ID", "Student Name", "Attendance"]
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1,column=col_num,value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        print("Records to be written to Excel: ", self.records)

        for row_num, (student_id,details) in enumerate(self.records.items(), start=2):
            sheet.cell(row=row_num, column=1, value=student_id)
            sheet.cell(row=row_num, column=2, value=details["name"])    
            sheet.cell(row=row_num, column=3, value=details["attendance"]) 

            wb.save("E:\CODING\Python (Basic to Advance)\Attendance_Record.xlsx")
            print("Excel file updated: E:\CODING\Python (Basic to Advance)\Attendance_Record.xlsx")

def main():
    attendance = AttendanceRecord()
        
    menu = """
        Attendance Management System
        1. Add Record
        2. View Record
        3. Update Record
        4. Replace Student ID
        5. Delete Record
        6. Exit
        """
    while True:
            print(menu)
            choice = input("Enter your choice (1-6): ")

            if choice == "1":
                student_id = input("Enter Student ID: ")
                student_name = input("Enter Student Name: ")
                attendance_status = input("Enter the Attendance status: ") 
                attendance.add_records(student_id,student_name,attendance_status)
                attendance.update_excel()

            elif choice == "2":
                attendance.view_records()
                attendance.update_excel()

            elif choice == "3":
                student_id = input("Enter Student ID to update: ")
                student_name = input("Enter the updated Student Name: ").strip()
                attendance_status = input("Enter the updated Attendance status (Present/Absent): ").strip()
                attendance.update_records(student_id, student_name, attendance_status)
                attendance.update_excel()

            elif choice == "4":
                old_id = input("Enter current Student ID to replace: ")
                new_id = input("Enter new Student ID: ")
                attendance.replace_student_id(old_id,new_id)
                attendance.update_excel()

            elif choice == "5":
                student_id = input("Enter Student ID: ")
                attendance.delete_records(student_id)
                attendance.update_excel()

            elif choice == "6":
                print("Exiting the program.")
                break

            else:
                print("invalid choice! Please try again.")

if __name__ == "__main__":
    main()
